# -*- coding: utf-8 -*-
"""
文件上传/下载路由
================

处理用户文件的上传、下载、列表查看和内容读取。
包含文件类型白名单、文件名清理、路径遍历防护等安全机制。

优化记录：
- [速率限制] 上传接口 10次/分钟，防止文件上传滥用
- [大小限制] MAX_FILE_SIZE = 20MB，防止超大文件占用服务器资源
"""

import os
import uuid
import re
from fastapi import APIRouter, UploadFile, File, HTTPException, Request
from fastapi.responses import FileResponse
from api.deps import UPLOAD_DIR
from slowapi import Limiter
from slowapi.util import get_remote_address

# [优化] 速率限制器 — 防止文件上传滥用
limiter = Limiter(key_func=get_remote_address)

router = APIRouter()


def _sanitize_filename(filename: str) -> str:
    """安全清理文件名，防止路径遍历攻击，保留 Unicode 字母/数字、下划线、连字符和点"""
    if not filename:
        return ""
    # 移除路径分隔符和危险字符
    filename = re.sub(r'[\\/:\*?"<>|]', "", filename)
    # 移除 .. 序列防止目录穿越
    filename = filename.replace("..", "")
    # 保留 Unicode 字母/数字 + 安全 ASCII 符号（支持中文等多语言文件名）
    filename = re.sub(r'[^\w.\-]', "", filename, flags=re.UNICODE)
    return filename.strip()


def _is_safe_path(base_dir: str, target_path: str) -> bool:
    """检查目标路径是否在安全目录内，防止路径遍历"""
    try:
        base_dir = os.path.abspath(base_dir)
        target_path = os.path.abspath(target_path)
        return target_path.startswith(base_dir + os.sep) or target_path == base_dir
    except Exception:
        return False


# 允许上传的文件扩展名白名单
ALLOWED_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt",
    ".py", ".js", ".jsx", ".ts", ".tsx", ".json", ".md", ".yaml", ".yml",
    ".html", ".htm", ".css", ".scss", ".sass", ".less", ".xml", ".svg",
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico", ".tiff", ".tif",
    ".mp3", ".wav", ".mp4", ".avi", ".mov", ".webm", ".flv", ".mkv",
    ".vue", ".svelte", ".astro",
    ".go", ".rs", ".java", ".c", ".cpp", ".h", ".hpp", ".cs",
    ".rb", ".php", ".sql", ".r", ".m", ".swift", ".kt", ".dart",
    ".sh", ".bash", ".zsh", ".bat", ".ps1",
    ".toml", ".ini", ".cfg", ".conf", ".properties",
    ".zip", ".tar", ".gz", ".rar", ".7z",
    ".log", ".map",
    ".woff", ".woff2", ".ttf", ".eot",
}

# 无扩展名但允许上传的文件名
NO_EXT_FILENAMES = {
    "makefile", "dockerfile", "vagrantfile", "gemfile", "rakefile",
    "procfile", "license", "readme", "changelog", "authors", "contributors",
}

# 点开头的特殊配置文件名
DOT_FILENAMES = {
    ".gitignore", ".editorconfig", ".eslintrc", ".prettierrc",
    ".babelrc", ".npmrc", ".zshrc", ".bashrc", ".bash_profile", ".profile",
    ".gitattributes", ".gitmodules",
}

# 最大文件大小限制：20MB
MAX_FILE_SIZE = 20 * 1024 * 1024


def _is_allowed_file(filename: str) -> bool:
    """检查文件是否在白名单中（按扩展名、无扩展名文件名、点文件名判断）"""
    ext = os.path.splitext(filename)[1].lower()
    if ext in ALLOWED_EXTENSIONS:
        return True
    name_lower = filename.lower()
    if name_lower in NO_EXT_FILENAMES:
        return True
    if name_lower in DOT_FILENAMES:
        return True
    if not ext and name_lower not in NO_EXT_FILENAMES:
        return False
    return False


@router.post("/upload")
@limiter.limit("10/minute")
async def upload_file(request: Request, file: UploadFile = File(...)):
    """上传文件 — 校验类型和大小，保存到 uploads 目录，返回文件 ID 和访问 URL"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename")

    filename = file.filename
    ext = os.path.splitext(filename)[1].lower()

    # 统一使用 _is_allowed_file 进行白名单检查
    if not _is_allowed_file(filename):
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext or filename}")

    # 读取文件内容并检查大小限制
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 20MB)")

    # 生成唯一文件 ID（UUID 前 12 位 + 原始扩展名）
    file_id = f"{uuid.uuid4().hex[:12]}{ext}"
    file_path = os.path.join(UPLOAD_DIR, file_id)

    with open(file_path, "wb") as f:
        f.write(content)

    file_type = _detect_file_type(ext)

    return {
        "file_id": file_id,
        "filename": filename,
        "size": len(content),
        "type": file_type,
        "url": f"/uploads/{file_id}",
    }


@router.get("/list")
async def list_files():
    """列出所有已上传的文件，按创建时间倒序排列"""
    if not os.path.exists(UPLOAD_DIR):
        return {"files": []}
    files = []
    for f in os.listdir(UPLOAD_DIR):
        fp = os.path.join(UPLOAD_DIR, f)
        if not os.path.isfile(fp):
            continue
        ext = os.path.splitext(f)[1].lower()
        stat = os.stat(fp)
        files.append({
            "file_id": f,
            "filename": f,
            "size": stat.st_size,
            "type": _detect_file_type(ext),
            "url": f"/uploads/{f}",
            "created_at": stat.st_ctime,
        })
    # 按创建时间倒序排列（最新的在前面）
    files.sort(key=lambda x: x["created_at"], reverse=True)
    return {"files": files}


@router.get("/download/{file_id}")
async def download_file(file_id: str):
    """下载指定文件 — 含路径遍历防护"""
    # 清理文件名，防止路径遍历
    file_id = _sanitize_filename(file_id)
    if not file_id:
        raise HTTPException(status_code=400, detail="Invalid file_id")

    fp = os.path.join(UPLOAD_DIR, file_id)
    # 安全检查：确保路径在允许的目录内
    if not _is_safe_path(UPLOAD_DIR, fp):
        raise HTTPException(status_code=403, detail="Access denied")
    if not os.path.exists(fp) or not os.path.isfile(fp):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(fp, filename=file_id)


# TTS 音频输出目录
TTS_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data", "tts_output")

# 生成图片输出目录
IMAGES_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data", "generated_images")


@router.get("/tts/{filename}")
async def download_tts_file(filename: str):
    """下载TTS生成的音频文件 — 含路径遍历防护"""
    # 清理文件名，防止路径遍历
    filename = _sanitize_filename(filename)
    if not filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    # 只允许 .mp3 扩展名
    if not filename.endswith(".mp3"):
        raise HTTPException(status_code=400, detail="Invalid file type")

    fp = os.path.join(TTS_OUTPUT_DIR, filename)
    # 安全检查：确保路径在允许的目录内
    if not _is_safe_path(TTS_OUTPUT_DIR, fp):
        raise HTTPException(status_code=403, detail="Access denied")
    if not os.path.exists(fp) or not os.path.isfile(fp):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(fp, filename=filename, media_type="audio/mpeg")


@router.get("/images/{filename}")
async def download_generated_image(filename: str):
    """下载AI生成的图片文件 — 含路径遍历防护"""
    # 清理文件名，防止路径遍历
    filename = _sanitize_filename(filename)
    if not filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    # 只允许图片扩展名
    ext = os.path.splitext(filename)[1].lower()
    if ext not in {".png", ".jpg", ".jpeg", ".webp", ".gif"}:
        raise HTTPException(status_code=400, detail="Invalid file type")

    fp = os.path.join(IMAGES_OUTPUT_DIR, filename)
    # 安全检查：确保路径在允许的目录内
    if not _is_safe_path(IMAGES_OUTPUT_DIR, fp):
        raise HTTPException(status_code=403, detail="Access denied")
    if not os.path.exists(fp) or not os.path.isfile(fp):
        raise HTTPException(status_code=404, detail="File not found")
    
    media_types = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".webp": "image/webp", ".gif": "image/gif"}
    return FileResponse(fp, filename=filename, media_type=media_types.get(ext, "image/png"))


# 文本类扩展名集合（用于文件内容读取接口）
TEXT_EXTENSIONS = {
    ".txt", ".md", ".log",
    ".py", ".js", ".jsx", ".ts", ".tsx", ".vue", ".svelte", ".astro",
    ".html", ".htm", ".css", ".scss", ".sass", ".less", ".xml", ".svg",
    ".json", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf", ".properties",
    ".go", ".rs", ".java", ".c", ".cpp", ".h", ".hpp", ".cs",
    ".rb", ".php", ".sql", ".r", ".m", ".swift", ".kt", ".dart",
    ".sh", ".bash", ".zsh", ".bat", ".ps1",
    ".map",
}


@router.get("/content/{file_id}")
async def get_file_content(file_id: str):
    """读取文件文本内容 — 支持 CSV、Excel、文本文件，二进制文件返回类型标识"""
    # 清理文件名，防止路径遍历
    file_id = _sanitize_filename(file_id)
    if not file_id:
        raise HTTPException(status_code=400, detail="Invalid file_id")

    fp = os.path.join(UPLOAD_DIR, file_id)
    # 安全检查：确保路径在允许的目录内
    if not _is_safe_path(UPLOAD_DIR, fp):
        raise HTTPException(status_code=403, detail="Access denied")
    if not os.path.exists(fp) or not os.path.isfile(fp):
        raise HTTPException(status_code=404, detail="File not found")

    ext = os.path.splitext(file_id)[1].lower()

    # CSV 文件直接读取文本内容
    if ext == ".csv":
        try:
            with open(fp, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
            return {"type": "csv", "content": content, "filename": file_id, "ext": ext}
        except Exception:
            raise HTTPException(status_code=500, detail="文件读取失败，请重新上传。")

    # Excel 文件使用 openpyxl 解析后转为文本格式返回
    elif ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(fp, read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
                rows.append("")  # 工作表之间加空行分隔
            wb.close()
            return {"type": "csv", "content": "\n".join(rows), "filename": file_id, "ext": ext}
        except Exception:
            raise HTTPException(status_code=500, detail="Excel 文件解析失败，请检查文件格式。")

    # 文本类文件直接读取
    elif ext in TEXT_EXTENSIONS:
        try:
            with open(fp, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
            return {"type": "text", "content": content, "filename": file_id, "ext": ext}
        except Exception:
            raise HTTPException(status_code=500, detail="文件读取失败，请重新上传。")

    # 其他类型标记为二进制，不返回内容
    else:
        return {"type": "binary", "content": None, "filename": file_id, "ext": ext}


# 文件类型检测常量（模块级，避免每次调用重建）
_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico", ".tiff", ".tif", ".svg"}
_DOC_EXTS = {".pdf", ".doc", ".docx"}
_SHEET_EXTS = {".xls", ".xlsx", ".csv"}
_AUDIO_EXTS = {".mp3", ".wav"}
_VIDEO_EXTS = {".mp4", ".avi", ".mov", ".webm", ".flv", ".mkv"}


def _detect_file_type(ext: str) -> str:
    """根据扩展名判断文件大类：image/document/spreadsheet/audio/video/text"""
    if ext in _IMAGE_EXTS:
        return "image"
    elif ext in _DOC_EXTS:
        return "document"
    elif ext in _SHEET_EXTS:
        return "spreadsheet"
    elif ext in _AUDIO_EXTS:
        return "audio"
    elif ext in _VIDEO_EXTS:
        return "video"
    else:
        return "text"
